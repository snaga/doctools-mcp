import openpyxl
import csv
import io
import os
import win32com.client
from doctools.pdf_service import export_pages_to_images as pdf_to_images
from doctools.util_service import format_error_response

def list_sheets(input_path: str) -> dict:
    """
    Get list of sheets in an Excel file.
    
    Args:
        input_path: Path to the XLSX file.
        
    Returns:
        dict: Result containing sheet names.
    """
    workbook = None
    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
            
        workbook = openpyxl.load_workbook(input_path, read_only=True)
        return {
            "status": "success",
            "sheets": workbook.sheetnames
        }
        
    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            workbook.close()

def extract_csv(input_path: str, output_dir: str = None, sheet_names: list[str] = None, encoding: str = "shift_jis") -> dict:
    """
    Extract specified Excel sheets (or all) as individual CSV files.
    
    Args:
        input_path: Path to the XLSX file.
        output_dir: Directory to save the CSV files (optional). 
                    Defaults to input file's directory.
        sheet_names: List of sheet names to extract (optional). 
                     Defaults to all sheets.
        encoding: Encoding for the output CSV files (optional, defaults to 'shift_jis').
        
    Returns:
        dict: Result containing list of output file paths and encoding.
    """
    workbook = None
    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
            
        workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        
        # Determine sheets to process
        if sheet_names is None:
            target_sheets = workbook.sheetnames
        else:
            # Handle single string if passed by mistake
            if isinstance(sheet_names, str):
                target_sheets = [sheet_names]
            else:
                target_sheets = sheet_names
                
        # Validate sheets
        for name in target_sheets:
            if name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{name}' not found.")
        
        # Determine output directory
        if output_dir is None:
            output_dir = os.path.dirname(os.path.abspath(input_path))
        
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_paths = []
        
        for sheet_name in target_sheets:
            filename = f"{base_name}_{sheet_name}.csv"
            path = os.path.join(output_dir, filename)
            
            sheet = workbook[sheet_name]
            with open(path, "w", encoding=encoding, newline="") as f:
                writer = csv.writer(f)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
            output_paths.append(path)
            
        return {
            "status": "success",
            "output_paths": output_paths,
            "encoding": encoding
        }
        
    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            workbook.close()

def extract_markdown_to_file(input_path: str, output_path: str = None, sheet_name: str = None) -> dict:
    """
    Extract data from an Excel sheet and save it as a Markdown table file.
    
    Args:
        input_path: Path to the XLSX file.
        output_path: Path to the output Markdown file (optional).
        sheet_name: Name of the sheet to extract (optional, defaults to first sheet).
        
    Returns:
        dict: Result containing the absolute path to the generated Markdown file.
    """
    workbook = None
    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
            
        workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        if sheet_name is None:
            sheet_name = workbook.sheetnames[0]
            
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found.")
            
        sheet = workbook[sheet_name]
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            content = ""
        else:
            md_lines = []
            
            # Header (first row)
            header = rows[0]
            header_cells = [str(cell).replace("\n", " ") if cell is not None else "" for cell in header]
            md_lines.append("| " + " | ".join(header_cells) + " |")
            md_lines.append("| " + " | ".join(["---"] * len(header_cells)) + " |")
            
            # Body
            for row in rows[1:]:
                # Pad row if shorter than header
                row_cells = []
                for i in range(len(header_cells)):
                    cell = row[i] if i < len(row) else None
                    cell_str = str(cell).replace("\n", " ") if cell is not None else ""
                    row_cells.append(cell_str)
                md_lines.append("| " + " | ".join(row_cells) + " |")
            content = "\n".join(md_lines)
            
        if output_path is None:
            output_path = os.path.splitext(input_path)[0] + ".md"
            
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
            
        return {
            "status": "success",
            "output_path": os.path.abspath(output_path)
        }
        
    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            workbook.close()

def extract_sheet(input_path: str, output_path: str, sheet_name: str) -> dict:
    """
    Extract a specific sheet from an Excel file and save as a new XLSX file.
    
    Args:
        input_path: Path to the source XLSX file.
        output_path: Path to save the new XLSX file.
        sheet_name: Name of the sheet to extract.
        
    Returns:
        dict: Result of the operation.
    """
    workbook = None
    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
            
        workbook = openpyxl.load_workbook(input_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found.")
            
        # Remove all other sheets
        for s_name in workbook.sheetnames:
            if s_name != sheet_name:
                del workbook[s_name]
                
        workbook.save(output_path)
        
        return {
            "status": "success",
            "message": f"Sheet '{sheet_name}' extracted successfully.",
            "output_path": output_path
        }
        
    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            workbook.close()

def export_sheets_to_images(
    input_path: str,
    output_dir: str = None,
    sheet_names: list[str] = None,
    dpi: int = 150
) -> dict:
    """
    Export specified Excel sheets as images (PNG/JPG) using Excel's PDF export and PyMuPDF.
    Requires Windows and Microsoft Excel installed.

    Args:
        input_path (str): Path to the source Excel file.
        output_dir (str, optional): Directory to save the images.
            If None, a directory based on the input filename is created.
        sheet_names (list[str], optional): List of sheet names to export.
            If None, only the active sheet is exported.
        dpi (int, optional): Resolution for the output images. Defaults to 150.

    Returns:
        dict: Result of the operation including status and list of output paths.
            Example: {"status": "success", "output_paths": ["path/to/sheet1_page_1.png", ...]}
            On failure: {"status": "error", "detail": "..."}
    """
    excel = None
    workbook = None
    temp_pdfs = []
    all_output_paths = []

    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")

        abs_input_path = os.path.abspath(input_path)
        
        # Determine output directory
        if output_dir is None:
            base_name = os.path.splitext(os.path.basename(input_path))[0]
            output_dir = os.path.join(os.path.dirname(abs_input_path), f"{base_name}_images")
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        # Initialize Excel
        # Use Dispatch to ensure we can control the instance
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(abs_input_path)

        # Determine sheets to process
        if sheet_names is None:
            target_sheets = [workbook.ActiveSheet.Name]
        else:
            # Handle single string if passed by mistake
            if isinstance(sheet_names, str):
                target_sheets = [sheet_names]
            else:
                target_sheets = sheet_names

        for sheet_name in target_sheets:
            try:
                sheet = workbook.Sheets(sheet_name)
            except Exception:
                raise ValueError(f"Sheet '{sheet_name}' not found in {input_path}")

            # Fit-to-Page Hack
            # 1. Set Zoom to False to enable FitToPages properties
            sheet.PageSetup.Zoom = False
            # 2. Fit to 1 page wide
            sheet.PageSetup.FitToPagesWide = 1
            # 3. Allow as many pages tall as needed (False or a large number)
            sheet.PageSetup.FitToPagesTall = False

            # Export to temporary PDF
            temp_pdf_path = os.path.join(output_dir, f"temp_{sheet_name}.pdf")
            # xlTypePDF = 0
            # Export the specific sheet
            sheet.ExportAsFixedFormat(0, temp_pdf_path)
            temp_pdfs.append(temp_pdf_path)

            # Convert PDF to images using pdf_service
            # We create a sub-directory for each sheet to avoid filename collisions
            sheet_output_dir = os.path.join(output_dir, sheet_name)
            res = pdf_to_images(temp_pdf_path, output_dir=sheet_output_dir, dpi=dpi)
            
            if res["status"] == "success":
                all_output_paths.extend(res["output_paths"])
            else:
                raise Exception(f"Failed to convert PDF to images for sheet '{sheet_name}': {res.get('detail')}")

        return {
            "status": "success",
            "output_paths": all_output_paths
        }

    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        
        # Clean up temporary PDFs
        for pdf_path in temp_pdfs:
            if os.path.exists(pdf_path):
                try:
                    os.remove(pdf_path)
                except Exception:
                    pass

def extract_structure(input_path: str) -> dict:
    """
    Extract the physical structure of an Excel file for PageIndex.
    Iterates through all sheets and samples the first few rows.

    Args:
        input_path (str): Path to the XLSX file.

    Returns:
        dict: Result containing the structure list.
            Each element: {"level": int, "title": str, "page": int, "sample_text": str}
    """
    workbook = None
    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
            
        workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        structure = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Sample first 10 rows
            rows_data = []
            for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                # Only include non-empty values
                row_str = " ".join([str(cell).strip() for cell in row if cell is not None])
                if row_str:
                    rows_data.append(row_str)
            
            sample = " | ".join(rows_data)[:500].strip()
            
            structure.append({
                "level": 1,
                "title": sheet_name,
                "page": 0, # Sheet-based, not page-based
                "sample_text": sample
            })
            
        return {
            "status": "success",
            "structure": structure
        }
    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            workbook.close()

def extract_node_content(input_path: str, sheet_name: str = None) -> dict:
    """
    Extract text content from a specified sheet of an Excel file.

    Args:
        input_path (str): Path to the XLSX file.
        sheet_name (str, optional): Name of the sheet.
            If None, all sheets are extracted.

    Returns:
        dict: Result containing the extracted text content.
    """
    workbook = None
    try:
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"Input file not found: {input_path}")
            
        workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        
        if sheet_name is None:
            target_sheets = workbook.sheetnames
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found.")
            target_sheets = [sheet_name]
            
        all_content = []
        for s_name in target_sheets:
            sheet = workbook[s_name]
            sheet_lines = [f"--- Sheet: {s_name} ---"]
            
            for row in sheet.iter_rows(values_only=True):
                # Format row as a pipe-separated string
                row_str = " | ".join([str(cell).replace("\n", " ").strip() if cell is not None else "" for cell in row])
                if row_str.strip().replace("|", "").strip(): # Only add non-empty rows
                    sheet_lines.append(row_str)
            
            all_content.append("\n".join(sheet_lines))
            
        return {
            "status": "success",
            "content": "\n\n".join(all_content)
        }
        
    except Exception as e:
        return format_error_response(e)
    finally:
        if workbook:
            workbook.close()
